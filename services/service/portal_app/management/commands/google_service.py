from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.utils import get_column_letter
from asgiref.sync import sync_to_async
from fpdf import FPDF
import os
from urllib.parse import quote_plus

SERVICE_ACCOUNT_FILE = "D:/diploma/bots/services/service/portal_app/creds/handy-woodland-452812-u0-d7eb0d4eb68d.json"
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

def get_drive_files_with_links(page_size=10):
    """
    Возвращает список словарей вида:
    [
      {
        "id": <str>,
        "name": <str>,
        "mimeType": <str>,
        "file_link": <str или None>
      },
      ...
    ]
    file_link — лучшая ссылка для открытия/редактирования/скачивания.
    Если файл не расшарен и нет webViewLink/webContentLink, может быть None.
    """
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build("drive", "v3", credentials=creds)

    results = service.files().list(
        pageSize=page_size,
        fields="files(id, name, mimeType, webViewLink, webContentLink)"
    ).execute()

    files = results.get('files', [])
    final_list = []
    for f in files:
        file_id = f.get("id")
        name = f.get("name")
        mime_type = f.get("mimeType")
        web_view = f.get("webViewLink")
        web_content = f.get("webContentLink")
        if web_view:
            file_link = web_view
        else:
            if mime_type == "application/vnd.google-apps.document":
                file_link = f"https://docs.google.com/document/d/{file_id}/edit"
            elif mime_type == "application/vnd.google-apps.spreadsheet":
                file_link = f"https://docs.google.com/spreadsheets/d/{file_id}/edit"
            else:
                file_link = web_content
        final_list.append({
            "id": file_id,
            "name": name,
            "mimeType": mime_type,
            "file_link": file_link
        })
    return final_list

def get_drive_files_by_folder(page_size=10, folder_id: str = None):
    """
    Возвращает список файлов из Google Drive, расположенных в папке с заданным folder_id.
    Если folder_id не указан или не найден, возвращается пустой список.
    Функция фильтрует файлы так, чтобы возвращались только те элементы, которые не являются папками.
    """
    if not folder_id:
        print("[DEBUG] folder_id не задан.", flush=True)
        return []
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build("drive", "v3", credentials=creds)
    query = f"'{folder_id}' in parents and trashed = false"
    print(f"[DEBUG] Запрос: {query}", flush=True)

    results = service.files().list(
         q=query,
         pageSize=page_size,
         fields="files(id, name, mimeType, parents, webViewLink, webContentLink)"
    ).execute()

    files = results.get('files', [])
    print(f"[DEBUG] Найдено файлов: {len(files)}", flush=True)

    final_list = []
    for f in files:
        # Если элемент является папкой, пропускаем его
        if f.get("mimeType") == "application/vnd.google-apps.folder":
            print(f"[DEBUG] Пропуск файла '{f.get('name')}', MIME тип - папка.", flush=True)
            continue

        # Дополнительная проверка: убеждаемся, что указанная папка входит в список родителей
        parents = f.get("parents", [])
        if folder_id not in parents:
            print(f"[DEBUG] Пропуск файла '{f.get('name')}', папки: {parents}", flush=True)
            continue

        file_id = f.get("id")
        name = f.get("name")
        mime_type = f.get("mimeType")
        web_view = f.get("webViewLink")
        web_content = f.get("webContentLink")
        if web_view:
            file_link = web_view
        else:
            if mime_type == "application/vnd.google-apps.document":
                file_link = f"https://docs.google.com/document/d/{file_id}/edit"
            elif mime_type in ["application/vnd.google-apps.spreadsheet", "application/vnd.google-apps.spreadsheets"]:
                file_link = f"https://docs.google.com/spreadsheets/d/{file_id}/edit"
            else:
                file_link = web_content
        final_list.append({
            "id": file_id,
            "name": name,
            "mimeType": mime_type,
            "file_link": file_link
        })
    return final_list

def get_salary_by_iin(iin: str):
    """
    Ищет в Google Sheets (название "Зарплаты") header row с ячейками "ИИН", "ФИО", "Зарплата",
    затем ниже этой строки ищет конкретный iin и возвращает (fio, salary). Если не найдено — возвращает None.
    """
    creds = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, SCOPES)
    client = gspread.authorize(creds)
    sheet = client.open("Зарплаты").sheet1
    data = sheet.get_all_values()
    if not data:
        print("DEBUG: Пустая таблица.")
        return None
    needed_headers = ["ИИН", "ФИО", "Зарплата"]
    header_row_index = None
    col_map = {}
    for i, row in enumerate(data):
        found_cols = {}
        for header in needed_headers:
            try:
                col_idx = row.index(header)
                found_cols[header] = col_idx
            except ValueError:
                break
        if len(found_cols) == len(needed_headers):
            header_row_index = i
            col_map = found_cols
            break
    if header_row_index is None:
        print("DEBUG: Не нашли строку, где есть все заголовки:", needed_headers)
        return None
    print(f"DEBUG: Заголовки найдены в строке {header_row_index+1}.")
    print("DEBUG: col_map =", col_map)
    iin_col = col_map["ИИН"]
    fio_col = col_map["ФИО"]
    salary_col = col_map["Зарплата"]
    for row_index in range(header_row_index+1, len(data)):
        row = data[row_index]
        if len(row) <= max(iin_col, fio_col, salary_col):
            continue
        row_iin = row[iin_col].strip()
        if row_iin == iin.strip():
            fio = row[fio_col]
            salary = row[salary_col]
            print(f"DEBUG: Найден iin={iin} в строке {row_index+1}: fio={fio}, salary={salary}")
            return (fio, salary)
    print(f"DEBUG: iin={iin} не найден ниже строки заголовков.")
    return None

def make_short_name_no_dots_for_user(user_obj) -> str:
    """
    Собирает user_obj.name + user_obj.first_name и превращает "Иванов Иван" → "ИвановИ" 
    (убирает точки и пробелы). Если одно из полей пустое, возвращает то, что есть.
    """
    surname = (user_obj.name or "").strip()
    fname = (user_obj.first_name or "").strip()
    if not surname and not fname:
        return ""
    full_name = f"{surname} {fname}".strip()
    parts = full_name.split()
    if len(parts) < 2:
        return full_name.replace(".", "").replace(" ", "")
    short = f"{parts[0]}{parts[1][0]}"
    short = short.replace(".", "").replace(" ", "")
    return short

def get_vacation_by_user_and_job(user_obj, job: str):
    """
    Формирует сокращённое ФИО из user_obj.name + user_obj.first_name, затем ищет
    в Google Sheets (название "График отпусков") строку, где:
      - Столбец "Ф.И.О." (без точек и пробелов, в нижнем регистре) начинается со сформированного сокращения.
      - Столбец "Должность" (в нижнем регистре) равен job.lower().
    Если найдено, возвращает словарь с ключами: days, agreed, transfer, note. Иначе — None.
    """
    short_fio = make_short_name_no_dots_for_user(user_obj)
    creds = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, SCOPES)
    client = gspread.authorize(creds)
    sheet = client.open("График отпусков").sheet1
    data = sheet.get_all_values()
    if not data:
        return None
    required_headers = [
        "ф.и.о.",
        "должность",
        "количество календарных дней",
        "согласованные дни отпуска",
        "перенесение отпуска",
        "примечание"
    ]
    header_row = None
    header_index = None
    for i, row in enumerate(data):
        row_clean = [cell.strip().lower() for cell in row]
        print(f"Строка {i}: {row_clean}")
        if all(header in row_clean for header in required_headers):
            header_row = row_clean
            header_index = i
            break
    if not header_row:
        print("Не найдены нужные заголовки в таблице.")
        return None
    try:
        fio_col = header_row.index("ф.и.о.")
        job_col = header_row.index("должность")
        days_col = header_row.index("количество календарных дней")
        agreed_col = header_row.index("согласованные дни отпуска")
        transfer_col = header_row.index("перенесение отпуска")
        note_col = header_row.index("примечание")
    except ValueError:
        print("Не найдены нужные заголовки в найденной строке:", header_row)
        return None
    short_fio_lower = short_fio.lower()
    job_lower = job.lower()
    for row in data[header_index + 1:]:
        if len(row) <= max(fio_col, job_col, days_col, agreed_col, transfer_col, note_col):
            continue
        row_fio = row[fio_col].replace(".", "").replace(" ", "").lower()
        row_job = row[job_col].strip().lower()
        if row_fio.startswith(short_fio_lower) and row_job == job_lower:
            return {
                "days": row[days_col],
                "agreed": row[agreed_col],
                "transfer": row[transfer_col],
                "note": row[note_col]
            }
    return None

def get_full_name(user_obj) -> str:
    return f"{user_obj.name or ''} {user_obj.first_name or ''}".strip()

def get_payroll_by_user_from_google_sheet(user_obj, month: str = None):
    """
    Ищет расчетный лист пользователя в Google Sheets (Таблица Зарплат).
    Функция ищет строку с заголовками, содержащими "ийн", затем ниже ищет конкретный iin
    и, если month задан, фильтрует по месяцу. Возвращает список словарей с данными записей.
    """
    creds = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, SCOPES)
    client = gspread.authorize(creds)
    sheet = client.open("Таблица Зарплат").sheet1
    data = sheet.get_all_values()
    if not data:
        return []
    header_row = None
    header_row_index = None
    for i in range(min(5, len(data))):
        row = data[i]
        if any("иин" in str(cell).strip().lower() for cell in row):
            header_row = [str(cell).strip() for cell in row]
            header_row_index = i
            break
    if header_row is None:
        print("DEBUG: Заголовок с 'ийн' не найден.")
        return []
    iin_idx = None
    for idx, h in enumerate(header_row):
        h_norm = h.replace(" ", "").lower()
        if h_norm == "иин" or "иин" in h_norm:
            iin_idx = idx
            break
    if iin_idx is None:
        print("DEBUG: Столбец с ИИН не найден.", header_row)
        return []
    month_idx = None
    for idx, h in enumerate(header_row):
        if "месяц" in h.lower():
            month_idx = idx
            break
    if month is not None and month_idx is None:
        print("DEBUG: Столбец с Месяцем не найден.")
    results = []
    user_iin = str(user_obj.iin or "").strip().lower()
    print("DEBUG: ИИН пользователя из БД:", user_iin)
    data_rows = data[header_row_index + 1:]
    for row in data_rows:
        if len(row) < len(header_row):
            continue
        row_norm = [str(cell).strip() for cell in row]
        row_iin = row_norm[iin_idx].lower()
        print("DEBUG: Сравниваем с ИИН из таблицы:", row_iin)
        if row_iin == user_iin:
            if month and month_idx is not None:
                row_month = row_norm[month_idx]
                if row_month != month:
                    continue
            payroll = dict(zip(header_row, row_norm))
            results.append(payroll)
    return results

def format_payroll(payroll: dict) -> str:
    """
    Форматирует словарь с данными расчетного листа для вывода.
    """
    msg = (
        f"ФИО: {payroll.get('ФИО')}\n"
        f"ИИН: {payroll.get('ИИН')}\n"
        f"Табельный номер: {payroll.get('Табельный номер')}\n"
        f"Должность: {payroll.get('Должность')}\n"
        f"Месяц: {payroll.get('Месяц')}\n"
        f"Оклад: {payroll.get('Оклад')}\n"
        f"Премия: {payroll.get('Премия')}\n"
        f"ИПН: {payroll.get('ИПН')}\n"
        f"ОПВ: {payroll.get('ОПВ')}\n"
        f"ОСМС: {payroll.get('ОСМС')}\n"
        f"Удержания: {payroll.get('Удержания')}\n"
        f"Итого к выплате: {payroll.get('Итого к выплате')}"
    )
    return msg

def save_payroll_to_pdf(payroll: dict, output_dir="/mnt/data/payroll_pdfs/") -> str:
    """
    Сохраняет расчетный лист (в виде словаря с данными) в PDF-файл.
    Файл будет сохранён в директории output_dir с именем, основанным на ФИО и месяце.
    Возвращает полный путь к PDF.
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    employee = payroll.get("ФИО", "unknown").replace(" ", "_")
    month = payroll.get("Месяц", "unknown").replace(" ", "_")
    filename = f"{employee}_{month}.pdf"
    output_path = os.path.join(output_dir, filename)
    
    pdf = FPDF()
    pdf.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
    pdf.add_page()
    pdf.set_font("DejaVu", size=12)
    pdf.cell(200, 10, txt="Расчетный лист / Есеп парағы", ln=True, align="C")
    pdf.ln(10)
    
    lines = format_payroll(payroll).split("\n")
    for line in lines:
        pdf.cell(200, 10, txt=line, ln=True)
    
    pdf.output(output_path)
    return output_path

if __name__ == "__main__":
    test_folder_id = "1cmcClsFSYmh9k3rW2JZtZS-e2C-KHeBf"
    files = get_drive_files_by_folder(page_size=10, folder_id=test_folder_id)
    print("Результат тестового вызова:", files, flush=True)